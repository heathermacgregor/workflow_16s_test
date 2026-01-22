"""
Result Export Utilities for Publication and Sharing

Functions to export analysis results in publication-ready formats:
- Excel workbooks with multiple sheets
- CSV files for supplementary data
- Summary tables in markdown/LaTeX
- Combined figures for publications

Author: GitHub Copilot (AI Assistant)
Date: 2026-01-08
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Union

import anndata as ad
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger('workflow_16s')


def export_results_to_excel(
    stats_df: pd.DataFrame,
    output_path: Union[str, Path],
    sheet_name: str = 'Statistical_Results',
    include_index: bool = False,
    freeze_panes: bool = True
) -> Path:
    """
    Export statistical results to Excel with formatting.
    
    Parameters
    ----------
    stats_df : pd.DataFrame
        Statistical results DataFrame
    output_path : str or Path
        Output Excel file path
    sheet_name : str, default='Statistical_Results'
        Name of the sheet
    include_index : bool, default=False
        Whether to include DataFrame index
    freeze_panes : bool, default=True
        Whether to freeze top row
    
    Returns
    -------
    Path
        Path to saved Excel file
    
    Examples
    --------
    >>> export_results_to_excel(
    ...     enhanced_stats,
    ...     'results.xlsx',
    ...     sheet_name='Differential_Abundance'
    ... )
    """
    output_path = Path(output_path)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(stats_df, index=include_index, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header row
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
    
    # Freeze top row
    if freeze_panes:
        ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    wb.save(output_path)
    logger.info(f"Exported results to {output_path}")
    
    return output_path


def export_publication_tables(
    stats_df: pd.DataFrame,
    output_dir: Union[str, Path],
    p_threshold: float = 0.05,
    effect_threshold: float = 0.33,
    top_n: int = 50
) -> Dict[str, Path]:
    """
    Export publication-ready tables.
    
    Creates:
    - Table 1: Top significant features (sorted by p-value and effect size)
    - Table 2: Summary statistics
    - Table S1: All significant features (supplementary)
    - Table S2: All features (supplementary)
    
    Parameters
    ----------
    stats_df : pd.DataFrame
        Statistical results with effect sizes
    output_dir : str or Path
        Output directory
    p_threshold : float, default=0.05
        FDR threshold for significance
    effect_threshold : float, default=0.33
        Effect size threshold (Cliff's delta)
    top_n : int, default=50
        Number of top features for main table
    
    Returns
    -------
    dict
        Paths to exported files
    
    Examples
    --------
    >>> tables = export_publication_tables(
    ...     enhanced_stats,
    ...     output_dir='publication_tables/',
    ...     top_n=20
    ... )
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    paths = {}
    
    # Determine significance columns
    p_col = 'p_adj' if 'p_adj' in stats_df.columns else 'p_value'
    effect_col = 'cliffs_delta' if 'cliffs_delta' in stats_df.columns else 'log2_fold_change'
    
    # Filter significant features
    significant = stats_df[stats_df[p_col] < p_threshold].copy()
    
    if effect_col in significant.columns:
        significant = significant[np.abs(significant[effect_col]) > effect_threshold]
    
    # Sort by significance
    significant = significant.sort_values([p_col, effect_col], ascending=[True, False])
    
    # Table 1: Top significant features
    table1 = significant.head(top_n)
    table1_path = output_dir / 'Table1_Top_Significant_Features.csv'
    table1.to_csv(table1_path, index=False)
    paths['table1'] = table1_path
    logger.info(f"Table 1 exported: {table1_path} ({len(table1)} features)")
    
    # Table 2: Summary statistics
    summary_data = {
        'Metric': [
            'Total features tested',
            'Significant features (FDR < 0.05)',
            f'Large effect (|{effect_col}| > {effect_threshold})',
            'Significant AND large effect',
            f'Top {top_n} features included in Table 1'
        ],
        'Count': [
            len(stats_df),
            (stats_df[p_col] < p_threshold).sum(),
            (np.abs(stats_df[effect_col]) > effect_threshold).sum() if effect_col in stats_df.columns else 0,
            len(significant),
            len(table1)
        ]
    }
    summary = pd.DataFrame(summary_data)
    summary_path = output_dir / 'Table2_Summary_Statistics.csv'
    summary.to_csv(summary_path, index=False)
    paths['table2'] = summary_path
    logger.info(f"Table 2 exported: {summary_path}")
    
    # Table S1: All significant features
    tableS1_path = output_dir / 'TableS1_All_Significant_Features.csv'
    significant.to_csv(tableS1_path, index=False)
    paths['tableS1'] = tableS1_path
    logger.info(f"Table S1 exported: {tableS1_path} ({len(significant)} features)")
    
    # Table S2: All features
    tableS2_path = output_dir / 'TableS2_All_Features.csv'
    stats_df.to_csv(tableS2_path, index=False)
    paths['tableS2'] = tableS2_path
    logger.info(f"Table S2 exported: {tableS2_path} ({len(stats_df)} features)")
    
    return paths


def export_supplementary_data(
    adata: ad.AnnData,
    output_dir: Union[str, Path],
    include_raw_counts: bool = True,
    include_metadata: bool = True,
    include_taxonomy: bool = True,
    include_tree: bool = False
) -> Dict[str, Path]:
    """
    Export supplementary data files for publication.
    
    Parameters
    ----------
    adata : ad.AnnData
        AnnData object
    output_dir : str or Path
        Output directory
    include_raw_counts : bool, default=True
        Export raw count matrix
    include_metadata : bool, default=True
        Export sample metadata
    include_taxonomy : bool, default=True
        Export feature taxonomy
    include_tree : bool, default=False
        Export phylogenetic tree (if available)
    
    Returns
    -------
    dict
        Paths to exported files
    
    Examples
    --------
    >>> export_supplementary_data(
    ...     adata,
    ...     output_dir='supplementary/',
    ...     include_raw_counts=True
    ... )
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    paths = {}
    
    # Export raw counts
    if include_raw_counts:
        counts_path = output_dir / 'SupplementaryData1_RawCounts.csv'
        
        # Convert to DataFrame
        if hasattr(adata.X, 'toarray'):
            counts_df = pd.DataFrame(
                adata.X.toarray(),
                index=adata.obs_names,
                columns=adata.var_names
            )
        else:
            counts_df = pd.DataFrame(
                adata.X,
                index=adata.obs_names,
                columns=adata.var_names
            )
        
        counts_df.to_csv(counts_path)
        paths['raw_counts'] = counts_path
        logger.info(f"Raw counts exported: {counts_path}")
    
    # Export metadata
    if include_metadata:
        metadata_path = output_dir / 'SupplementaryData2_SampleMetadata.csv'
        adata.obs.to_csv(metadata_path)
        paths['metadata'] = metadata_path
        logger.info(f"Metadata exported: {metadata_path}")
    
    # Export taxonomy
    if include_taxonomy:
        taxonomy_path = output_dir / 'SupplementaryData3_FeatureTaxonomy.csv'
        adata.var.to_csv(taxonomy_path)
        paths['taxonomy'] = taxonomy_path
        logger.info(f"Taxonomy exported: {taxonomy_path}")
    
    # Export tree (if available)
    if include_tree and 'phylogenetic_tree' in adata.uns:
        tree_path = output_dir / 'SupplementaryData4_PhylogeneticTree.nwk'
        with open(tree_path, 'w') as f:
            f.write(adata.uns['phylogenetic_tree'])
        paths['tree'] = tree_path
        logger.info(f"Phylogenetic tree exported: {tree_path}")
    
    return paths


def create_methods_section(
    adata: ad.AnnData,
    stats_config: Dict,
    output_path: Optional[Union[str, Path]] = None
) -> str:
    """
    Generate methods section text for publications.
    
    Parameters
    ----------
    adata : ad.AnnData
        AnnData object
    stats_config : dict
        Statistical analysis configuration
    output_path : str or Path, optional
        If provided, saves to file
    
    Returns
    -------
    str
        Methods section text
    
    Examples
    --------
    >>> methods = create_methods_section(
    ...     adata,
    ...     stats_config={'effect_sizes': True, 'permutation_tests': True}
    ... )
    >>> print(methods)
    """
    methods = []
    
    # Sample information
    methods.append(f"## Statistical Analysis\n")
    methods.append(f"A total of {adata.n_obs} samples and {adata.n_vars} features ")
    methods.append(f"were analyzed. ")
    
    # Effect sizes
    if stats_config.get('effect_sizes'):
        methods.append(
            "Effect sizes were calculated using Cliff's delta for non-parametric "
            "comparison and Cohen's d for standardized mean differences. "
        )
    
    # Batch correction
    if stats_config.get('batch_correction'):
        method = stats_config.get('batch_correction_method', 'percentile')
        methods.append(
            f"Batch effects were corrected using {method} normalization, "
            "an approach specifically designed for microbiome compositional data. "
        )
    
    # Rarefaction
    if stats_config.get('rarefaction'):
        methods.append(
            "Sequencing depth adequacy was assessed using rarefaction curves "
            "to ensure saturation of microbial diversity. "
        )
    
    # Permutation tests
    if stats_config.get('permutation_tests'):
        n_perms = stats_config.get('n_permutations', 10000)
        methods.append(
            f"Statistical significance was assessed using permutation tests "
            f"with {n_perms:,} permutations, controlling family-wise error rate (FWER) "
            "using the Max-T step-down procedure. "
        )
    
    # Decontam
    if stats_config.get('decontam'):
        methods.append(
            "Potential contaminants were identified using the decontam R package "
            "based on inverse correlation with DNA concentration and prevalence "
            "in negative control samples. "
        )
    
    # Combine
    methods_text = ''.join(methods)
    
    # Save if path provided
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            f.write(methods_text)
        logger.info(f"Methods section exported: {output_path}")
    
    return methods_text


def export_complete_results_package(
    adata: ad.AnnData,
    stats_df: pd.DataFrame,
    output_dir: Union[str, Path],
    stats_config: Optional[Dict] = None,
    p_threshold: float = 0.05,
    effect_threshold: float = 0.33
) -> Dict[str, Union[Path, Dict[str, Path]]]:
    """
    Export complete results package for publication.
    
    Creates organized directory structure with:
    - Main results Excel file
    - Publication tables (CSV)
    - Supplementary data
    - Methods section template
    
    Parameters
    ----------
    adata : ad.AnnData
        AnnData object
    stats_df : pd.DataFrame
        Statistical results
    output_dir : str or Path
        Output directory
    stats_config : dict, optional
        Statistical analysis configuration
    p_threshold : float, default=0.05
        FDR threshold
    effect_threshold : float, default=0.33
        Effect size threshold
    
    Returns
    -------
    dict
        Paths to all exported files
    
    Examples
    --------
    >>> package = export_complete_results_package(
    ...     adata,
    ...     enhanced_stats,
    ...     output_dir='publication_package/',
    ...     stats_config=config.statistical_analysis
    ... )
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"Exporting complete results package to {output_dir}")
    
    all_paths = {}
    
    # 1. Main Excel file
    excel_path = export_results_to_excel(
        stats_df,
        output_dir / 'Main_Results.xlsx'
    )
    all_paths['main_excel'] = excel_path
    
    # 2. Publication tables
    tables_dir = output_dir / 'publication_tables'
    table_paths = export_publication_tables(
        stats_df,
        tables_dir,
        p_threshold=p_threshold,
        effect_threshold=effect_threshold
    )
    all_paths['publication_tables'] = table_paths
    
    # 3. Supplementary data
    supp_dir = output_dir / 'supplementary_data'
    supp_paths = export_supplementary_data(adata, supp_dir)
    all_paths['supplementary_data'] = supp_paths
    
    # 4. Methods section
    if stats_config:
        methods_text = create_methods_section(
            adata,
            stats_config,
            output_path=output_dir / 'methods_section.md'
        )
        all_paths['methods'] = output_dir / 'methods_section.md'
    
    # 5. Create README
    readme_path = output_dir / 'README.md'
    readme_content = f"""# Analysis Results Package

Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

## Contents

### Main Results
- `Main_Results.xlsx` - All statistical results in Excel format

### Publication Tables
- `publication_tables/Table1_Top_Significant_Features.csv` - Top {len(pd.read_csv(table_paths['table1']))} significant features
- `publication_tables/Table2_Summary_Statistics.csv` - Summary of analysis
- `publication_tables/TableS1_All_Significant_Features.csv` - All significant features (supplementary)
- `publication_tables/TableS2_All_Features.csv` - Complete results (supplementary)

### Supplementary Data
- `supplementary_data/SupplementaryData1_RawCounts.csv` - Raw abundance matrix
- `supplementary_data/SupplementaryData2_SampleMetadata.csv` - Sample metadata
- `supplementary_data/SupplementaryData3_FeatureTaxonomy.csv` - Feature taxonomy

### Methods
- `methods_section.md` - Draft methods section for manuscript

## Dataset Summary
- Samples: {adata.n_obs:,}
- Features: {adata.n_vars:,}
- Significant features (FDR < {p_threshold}): {(stats_df['p_adj'] < p_threshold).sum() if 'p_adj' in stats_df.columns else 'N/A'}

## Citation
If you use these results, please cite the workflow_16s pipeline and relevant statistical methods.
"""
    
    with open(readme_path, 'w') as f:
        f.write(readme_content)
    all_paths['readme'] = readme_path
    
    logger.info(f"Complete results package exported to {output_dir}")
    logger.info(f"Total files created: {len(all_paths)}")
    
    return all_paths
